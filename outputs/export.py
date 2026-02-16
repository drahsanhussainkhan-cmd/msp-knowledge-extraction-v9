"""
Export utilities - CSV, JSON, and Excel export for extraction results.
"""

import csv
import json
import os
from collections import defaultdict


def export_to_csv(results, output_path):
    """Export all extractions to a flat CSV file."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    rows = []
    if isinstance(results, dict):
        for doc_name, doc_results in results.items():
            if isinstance(doc_results, dict):
                for category, items in doc_results.items():
                    if isinstance(items, list):
                        for item in items:
                            row = {"document": doc_name, "category": category}
                            if isinstance(item, dict):
                                row.update(item)
                            elif hasattr(item, "__dataclass_fields__"):
                                import dataclasses
                                row.update(dataclasses.asdict(item))
                            else:
                                row["value"] = str(item)
                            rows.append(row)

    if not rows:
        print("No data to export.")
        return None

    # Collect all field names
    fieldnames = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                fieldnames.append(key)
                seen.add(key)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    print(f"CSV exported to {output_path} ({len(rows)} rows)")
    return output_path


def export_to_json(results, output_path):
    """Export results to a structured JSON file."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)

    print(f"JSON exported to {output_path}")
    return output_path


def export_to_excel(results, output_path):
    """Export results to Excel with one sheet per category."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl is required for Excel export. Install with: pip install openpyxl")
        return None

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Group by category
    by_category = defaultdict(list)
    if isinstance(results, dict):
        for doc_name, doc_results in results.items():
            if isinstance(doc_results, dict):
                for category, items in doc_results.items():
                    if isinstance(items, list):
                        for item in items:
                            row = {"document": doc_name}
                            if isinstance(item, dict):
                                row.update(item)
                            elif hasattr(item, "__dataclass_fields__"):
                                import dataclasses
                                row.update(dataclasses.asdict(item))
                            else:
                                row["value"] = str(item)
                            by_category[category].append(row)

    if not by_category:
        # Create at least a summary sheet
        ws = wb.create_sheet("Summary")
        ws.append(["No extractions found"])
        wb.save(output_path)
        return output_path

    for category, rows in sorted(by_category.items()):
        # Excel sheet names max 31 chars
        sheet_name = category[:31]
        ws = wb.create_sheet(sheet_name)

        # Headers
        fieldnames = []
        seen = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    fieldnames.append(key)
                    seen.add(key)

        ws.append(fieldnames)

        # Data rows
        for row in rows:
            ws.append([str(row.get(f, "")) for f in fieldnames])

    wb.save(output_path)
    print(f"Excel exported to {output_path} ({len(by_category)} sheets)")
    return output_path


def export_gaps_to_csv(gaps, output_path):
    """Export gap analysis results to CSV."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    if not gaps:
        print("No gaps to export.")
        return None

    rows = []
    for gap in gaps:
        if hasattr(gap, "__dataclass_fields__"):
            import dataclasses
            rows.append(dataclasses.asdict(gap))
        elif isinstance(gap, dict):
            rows.append(gap)
        else:
            rows.append({"description": str(gap)})

    fieldnames = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                fieldnames.append(key)
                seen.add(key)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    print(f"Gaps CSV exported to {output_path} ({len(rows)} rows)")
    return output_path
